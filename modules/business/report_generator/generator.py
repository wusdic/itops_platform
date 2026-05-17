"""
报告生成器服务
支持PDF和Excel导出
"""

import os
import io
import uuid
from datetime import datetime
from typing import Optional, List, Dict, Any
import logging
from concurrent.futures import ThreadPoolExecutor

from jinja2 import Environment, BaseLoader, TemplateNotFound

logger = logging.getLogger(__name__)

# 尝试导入PDF和Excel库
try:
    from weasyprint import HTML
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False
    logger.warning("weasyprint not installed, PDF export will not be available")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed, Excel export will not be available")


class ReportDataCollector:
    """报告数据收集器基类"""
    
    def collect_device_status(self, start_date: datetime, end_date: datetime, 
                              filters: Dict = None) -> Dict[str, Any]:
        """收集设备状态数据"""
        return {
            "report_date": datetime.now(),
            "generated_by": "System",
            "devices": [
                {"name": "Server-01", "ip": "192.168.1.10", "status": "online", "cpu_usage": 45, "memory_usage": 60},
                {"name": "Server-02", "ip": "192.168.1.11", "status": "online", "cpu_usage": 32, "memory_usage": 45},
                {"name": "Server-03", "ip": "192.168.1.12", "status": "warning", "cpu_usage": 78, "memory_usage": 85},
            ],
            "total_devices": 3,
            "online_devices": 3,
            "offline_devices": 0,
            "summary": {
                "total_alerts": 5,
                "critical": 1,
                "warning": 2,
                "info": 2
            }
        }
    
    def collect_alert_summary(self, start_date: datetime, end_date: datetime,
                              filters: Dict = None) -> Dict[str, Any]:
        """收集告警汇总数据"""
        return {
            "report_date": datetime.now(),
            "generated_by": "System",
            "alerts": [
                {"level": "critical", "count": 3, "devices": ["Server-01", "Server-03"]},
                {"level": "warning", "count": 12, "devices": ["Server-02", "Server-04"]},
                {"level": "info", "count": 45, "devices": []},
            ],
            "total_alerts": 60,
            "top_alert_devices": [
                {"name": "Server-01", "count": 15},
                {"name": "Server-02", "count": 10},
            ]
        }
    
    def collect_performance_trend(self, start_date: datetime, end_date: datetime,
                                  filters: Dict = None) -> Dict[str, Any]:
        """收集性能趋势数据"""
        return {
            "report_date": datetime.now(),
            "generated_by": "System",
            "metrics": [
                {"name": "CPU Usage", "avg": 45.5, "max": 89.2, "min": 12.3},
                {"name": "Memory Usage", "avg": 62.3, "max": 85.1, "min": 30.5},
                {"name": "Disk Usage", "avg": 55.0, "max": 72.0, "min": 40.0},
            ],
            "trends": [
                {"date": "2024-01-01", "cpu": 42, "memory": 58},
                {"date": "2024-01-02", "cpu": 45, "memory": 60},
            ]
        }


class ReportExporter:
    """报告导出器"""
    
    def __init__(self, output_dir: str = "/tmp/reports"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def export_pdf(self, html_content: str, filename: str) -> str:
        """导出PDF"""
        if not WEASYPRINT_AVAILABLE:
            raise ImportError("weasyprint is not installed")

        file_path = os.path.join(self.output_dir, filename)
        HTML(string=html_content).write_pdf(file_path)
        return file_path

    def export_html(self, html_content: str, filename: str) -> str:
        """导出HTML"""
        file_path = os.path.join(self.output_dir, filename)
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        return file_path

    def export_word(self, html_content: str, filename: str) -> str:
        """导出Word (HTML格式)"""
        file_path = os.path.join(self.output_dir, filename.replace('.docx', '.html'))
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        return file_path

    def export_markdown(self, content: str, filename: str) -> str:
        """导出Markdown"""
        file_path = os.path.join(self.output_dir, filename)
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(content)
        return file_path

    def export_json(self, data: Dict[str, Any], filename: str) -> str:
        """导出JSON"""
        import json
        file_path = os.path.join(self.output_dir, filename)
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return file_path

    def export_excel(self, data: Dict[str, Any], filename: str) -> str:
        """导出Excel"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is not installed")
        
        file_path = os.path.join(self.output_dir, filename)
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # 样式定义
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写入标题
        row = 1
        ws.cell(row=row, column=1, value=data.get("title", "Report"))
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 1
        
        # 写入报告信息
        ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        row += 2
        
        # 写入数据表
        headers = []
        if "data" in data and isinstance(data["data"], list) and len(data["data"]) > 0:
            headers = list(data["data"][0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            row += 1
            for record in data["data"]:
                for col, key in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=record.get(key, ""))
                    cell.border = border
                row += 1
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        wb.save(file_path)
        return file_path


class ReportGenerator:
    """报告生成引擎"""
    
    def __init__(self, output_dir: str = "/tmp/reports"):
        self.output_dir = output_dir
        self.jinja_env = Environment(loader=BaseLoader())
        self.data_collector = ReportDataCollector()
        self.exporter = ReportExporter(output_dir)
        self.executor = ThreadPoolExecutor(max_workers=4)
        os.makedirs(output_dir, exist_ok=True)
    
    def generate_report_no(self) -> str:
        """生成报告编号"""
        return f"RPT-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"
    
    def render_template(self, template_content: str, context: Dict[str, Any]) -> str:
        """渲染Jinja2模板"""
        try:
            template = self.jinja_env.from_string(template_content)
            return template.render(**context)
        except Exception as e:
            logger.error(f"Template render error: {e}")
            raise
    
    def collect_data(self, template_type: str, start_date: datetime,
                     end_date: datetime, filters: Dict = None) -> Dict[str, Any]:
        """收集报告数据"""
        collectors = {
            "device_status": self.data_collector.collect_device_status,
            "alert_summary": self.data_collector.collect_alert_summary,
            "performance_trend": self.data_collector.collect_performance_trend,
        }
        
        collector = collectors.get(template_type)
        if collector:
            return collector(start_date, end_date, filters)
        
        return {"report_date": datetime.now(), "generated_by": "System"}
    
    def generate(self, template_content: str, template_type: str,
                 start_date: datetime, end_date: datetime,
                 format: str = "pdf", filters: Dict = None,
                 params: Dict = None) -> Dict[str, Any]:
        """
        生成报告
        
        Args:
            template_content: Jinja2模板内容
            template_type: 报告类型
            start_date: 开始日期
            end_date: 结束日期
            format: 输出格式 (pdf/excel)
            filters: 筛选条件
            params: 额外参数
            
        Returns:
            包含报告信息的字典
        """
        # 收集数据
        data = self.collect_data(template_type, start_date, end_date, filters)
        data["start_date"] = start_date
        data["end_date"] = end_date
        if params:
            data.update(params)
        
        # 渲染模板
        html_content = self.render_template(template_content, data)
        
        # 生成报告编号和文件名
        report_no = self.generate_report_no()
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        
        # 导出文件
        if format == "pdf":
            filename = f"{report_no}.pdf"
            file_path = self.exporter.export_pdf(html_content, filename)
        elif format == "excel":
            filename = f"{report_no}.xlsx"
            file_path = self.exporter.export_excel(
                {"title": params.get("title", "Report") if params else "Report", "data": data.get("devices", [])},
                filename
            )
        else:
            raise ValueError(f"Unsupported format: {format}")
        
        # 获取文件大小
        file_size = os.path.getsize(file_path)
        
        return {
            "report_no": report_no,
            "file_path": file_path,
            "file_name": filename,
            "file_size": file_size,
            "format": format,
            "status": "completed",
            "generated_at": datetime.now()
        }
    
    def generate_async(self, template_content: str, template_type: str,
                      start_date: datetime, end_date: datetime,
                      format: str = "pdf", filters: Dict = None,
                      params: Dict = None, callback=None):
        """异步生成报告"""
        future = self.executor.submit(
            self.generate, template_content, template_type,
            start_date, end_date, format, filters, params
        )
        if callback:
            future.add_done_callback(callback)
        return future


# 内置模板
BUILTIN_TEMPLATES = {
    "device_status": """<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>设备状态报告</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 40px; }
        h1 { color: #333; border-bottom: 2px solid #4472C4; padding-bottom: 10px; }
        table { width: 100%; border-collapse: collapse; margin-top: 20px; }
        th, td { border: 1px solid #ddd; padding: 12px; text-align: left; }
        th { background-color: #4472C4; color: white; }
        tr:nth-child(even) { background-color: #f9f9f9; }
        .summary { background: #f0f7ff; padding: 15px; border-radius: 5px; margin: 20px 0; }
        .status-online { color: green; }
        .status-warning { color: orange; }
        .status-offline { color: red; }
    </style>
</head>
<body>
    <h1>设备状态报告</h1>
    <div class="summary">
        <p><strong>报告日期:</strong> {{ report_date.strftime('%Y-%m-%d %H:%M:%S') }}</p>
        <p><strong>报告周期:</strong> {{ start_date.strftime('%Y-%m-%d') }} ~ {{ end_date.strftime('%Y-%m-%d') }}</p>
        <p><strong>生成人:</strong> {{ generated_by }}</p>
    </div>
    
    <h2>概览</h2>
    <div class="summary">
        <p>设备总数: {{ total_devices }} | 在线: {{ online_devices }} | 离线: {{ offline_devices }}</p>
        <p>告警统计: 严重 {{ summary.critical }} | 警告 {{ summary.warning }} | 提示 {{ summary.info }}</p>
    </div>
    
    <h2>设备详情</h2>
    <table>
        <thead>
            <tr>
                <th>设备名称</th>
                <th>IP地址</th>
                <th>状态</th>
                <th>CPU使用率</th>
                <th>内存使用率</th>
            </tr>
        </thead>
        <tbody>
        {% for device in devices %}
            <tr>
                <td>{{ device.name }}</td>
                <td>{{ device.ip }}</td>
                <td class="status-{{ device.status }}">{{ device.status|upper }}</td>
                <td>{{ device.cpu_usage }}%</td>
                <td>{{ device.memory_usage }}%</td>
            </tr>
        {% endfor %}
        </tbody>
    </table>
</body>
</html>""",
    
    "alert_summary": """<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>告警汇总报告</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 40px; }
        h1 { color: #333; border-bottom: 2px solid #E74C3C; padding-bottom: 10px; }
        table { width: 100%; border-collapse: collapse; margin-top: 20px; }
        th, td { border: 1px solid #ddd; padding: 12px; text-align: left; }
        th { background-color: #E74C3C; color: white; }
        .level-critical { background-color: #FADBD8; }
        .level-warning { background-color: #FCF3CF; }
        .level-info { background-color: #D5F5E3; }
        .summary { background: #FDEDEC; padding: 15px; border-radius: 5px; margin: 20px 0; }
    </style>
</head>
<body>
    <h1>告警汇总报告</h1>
    <div class="summary">
        <p><strong>报告日期:</strong> {{ report_date.strftime('%Y-%m-%d %H:%M:%S') }}</p>
        <p><strong>报告周期:</strong> {{ start_date.strftime('%Y-%m-%d') }} ~ {{ end_date.strftime('%Y-%m-%d') }}</p>
        <p><strong>告警总数:</strong> {{ total_alerts }}</p>
    </div>
    
    <h2>告警级别统计</h2>
    <table>
        <thead>
            <tr>
                <th>级别</th>
                <th>数量</th>
                <th>涉及设备</th>
            </tr>
        </thead>
        <tbody>
        {% for alert in alerts %}
            <tr class="level-{{ alert.level }}">
                <td>{{ alert.level|upper }}</td>
                <td>{{ alert.count }}</td>
                <td>{{ alert.devices|join(', ') if alert.devices else '-' }}</td>
            </tr>
        {% endfor %}
        </tbody>
    </table>
    
    <h2>告警设备排行</h2>
    <table>
        <thead>
            <tr>
                <th>排名</th>
                <th>设备名称</th>
                <th>告警数量</th>
            </tr>
        </thead>
        <tbody>
        {% for item in top_alert_devices %}
            <tr>
                <td>{{ loop.index }}</td>
                <td>{{ item.name }}</td>
                <td>{{ item.count }}</td>
            </tr>
        {% endfor %}
        </tbody>
    </table>
</body>
</html>""",
    
    "performance_trend": """<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>性能趋势报告</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 40px; }
        h1 { color: #333; border-bottom: 2px solid #27AE60; padding-bottom: 10px; }
        table { width: 100%; border-collapse: collapse; margin-top: 20px; }
        th, td { border: 1px solid #ddd; padding: 12px; text-align: left; }
        th { background-color: #27AE60; color: white; }
        .metric-header { background-color: #E8F8F5; }
        .summary { background: #EAFAF1; padding: 15px; border-radius: 5px; margin: 20px 0; }
    </style>
</head>
<body>
    <h1>性能趋势报告</h1>
    <div class="summary">
        <p><strong>报告日期:</strong> {{ report_date.strftime('%Y-%m-%d %H:%M:%S') }}</p>
        <p><strong>报告周期:</strong> {{ start_date.strftime('%Y-%m-%d') }} ~ {{ end_date.strftime('%Y-%m-%d') }}</p>
    </div>
    
    <h2>性能指标汇总</h2>
    <table>
        <thead>
            <tr>
                <th>指标名称</th>
                <th>平均值</th>
                <th>最大值</th>
                <th>最小值</th>
            </tr>
        </thead>
        <tbody>
        {% for metric in metrics %}
            <tr class="metric-header">
                <td><strong>{{ metric.name }}</strong></td>
                <td>{{ "%.1f"|format(metric.avg) }}%</td>
                <td>{{ "%.1f"|format(metric.max) }}%</td>
                <td>{{ "%.1f"|format(metric.min) }}%</td>
            </tr>
        {% endfor %}
        </tbody>
    </table>
    
    <h2>趋势数据</h2>
    <table>
        <thead>
            <tr>
                <th>日期</th>
                <th>CPU使用率</th>
                <th>内存使用率</th>
            </tr>
        </thead>
        <tbody>
        {% for trend in trends %}
            <tr>
                <td>{{ trend.date }}</td>
                <td>{{ trend.cpu }}%</td>
                <td>{{ trend.memory }}%</td>
            </tr>
        {% endfor %}
        </tbody>
    </table>
</body>
</html>"""
}