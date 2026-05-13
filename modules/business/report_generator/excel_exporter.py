"""
Work Order Excel Exporter Module
Provides Excel export functionality for work orders
"""

import io
import logging
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime
from enum import Enum

logger = logging.getLogger(__name__)

# Try to import openpyxl, fall back to xlsxwriter
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_ENGINE = 'openpyxl'
except ImportError:
    try:
        import xlsxwriter
        EXCEL_ENGINE = 'xlsxwriter'
    except ImportError:
        EXCEL_ENGINE = None


class ExportFormat(str, Enum):
    """Export format options"""
    XLSX = "xlsx"
    CSV = "csv"


class WorkOrderExporter:
    """
    Work Order Excel Exporter
    
    Exports work orders to Excel/CSV format with proper formatting.
    Supports filtering by status, priority, and date range.
    """
    
    # Column definitions with widths
    COLUMNS = [
        {'key': 'order_no', 'header': '工单编号', 'width': 18},
        {'key': 'order_type', 'header': '工单类型', 'width': 12},
        {'key': 'title', 'header': '标题', 'width': 40},
        {'key': 'priority', 'header': '优先级', 'width': 8},
        {'key': 'status', 'header': '状态', 'width': 12},
        {'key': 'creator', 'header': '创建人', 'width': 15},
        {'key': 'assignee', 'header': '处理人', 'width': 15},
        {'key': 'device_name', 'header': '设备名称', 'width': 20},
        {'key': 'device_ip', 'header': '设备IP', 'width': 15},
        {'key': 'created_at', 'header': '创建时间', 'width': 20},
        {'key': 'updated_at', 'header': '更新时间', 'width': 20},
        {'key': 'expected_end', 'header': '期望完成', 'width': 20},
        {'key': 'actual_end', 'header': '实际完成', 'width': 20},
        {'key': 'sla_response_time', 'header': 'SLA响应(分钟)', 'width': 15},
        {'key': 'sla_resolve_time', 'header': 'SLA解决(分钟)', 'width': 15},
        {'key': 'sla_breached', 'header': 'SLA超时', 'width': 10},
        {'key': 'description', 'header': '描述', 'width': 50},
        {'key': 'resolution', 'header': '解决方案', 'width': 50},
        {'key': 'root_cause', 'header': '根本原因', 'width': 40},
        {'key': 'improvement', 'header': '改进措施', 'width': 40},
        {'key': 'impact', 'header': '影响范围', 'width': 15},
        {'key': 'tags', 'header': '标签', 'width': 30},
        {'key': 'closed_at', 'header': '关闭时间', 'width': 20},
    ]
    
    # Status to Chinese mapping
    STATUS_MAP = {
        'pending': '待处理',
        'processing': '处理中',
        'pending_approval': '待审批',
        'approved': '已批准',
        'rejected': '已拒绝',
        'resolved': '已解决',
        'closed': '已关闭',
        'cancelled': '已取消',
    }
    
    # Type to Chinese mapping
    TYPE_MAP = {
        'fault': '故障',
        'change': '变更',
        'inspection': '巡检',
        'security': '安全',
        'demand': '需求',
        'question': '咨询',
        'other': '其他',
    }
    
    # Priority to Chinese mapping
    PRIORITY_MAP = {
        'P1': 'P1-紧急',
        'P2': 'P2-高',
        'P3': 'P3-中',
        'P4': 'P4-低',
    }
    
    # Priority colors
    PRIORITY_COLORS = {
        'P1': 'FF0000',  # Red
        'P2': 'FF6600',  # Orange
        'P3': '0066CC',  # Blue
        'P4': '00CC00',  # Green
    }
    
    # Status colors
    STATUS_COLORS = {
        'pending': 'FFFF00',      # Yellow
        'processing': '00BFFF',   # Deep sky blue
        'pending_approval': 'FFA500',  # Orange
        'approved': '00FF00',    # Green
        'rejected': 'FF0000',    # Red
        'resolved': '90EE90',    # Light green
        'closed': 'C0C0C0',     # Silver
        'cancelled': '808080',   # Gray
    }
    
    def __init__(self):
        """Initialize exporter"""
        pass
    
    def export(
        self,
        workorders: List[Dict[str, Any]],
        filename: str = None,
        format: ExportFormat = ExportFormat.XLSX,
        include_columns: List[str] = None,
        title: str = '工单导出'
    ) -> Tuple[bytes, str]:
        """
        Export work orders to Excel format
        
        Args:
            workorders: List of work order dictionaries
            filename: Output filename (without extension)
            format: Export format (XLSX or CSV)
            include_columns: List of column keys to include (None for all)
            title: Sheet title
            
        Returns:
            (file_bytes, content_type)
        """
        if not filename:
            filename = f"workorders_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        if format == ExportFormat.CSV:
            return self._export_csv(workorders, filename, include_columns)
        else:
            return self._export_xlsx(workorders, filename, include_columns, title)
    
    def export_workorders(
        self,
        workorders: List[Dict[str, Any]],
        filters: Optional[Dict[str, Any]] = None,
        format: ExportFormat = ExportFormat.XLSX
    ) -> Tuple[bytes, str]:
        """
        Export work orders with optional filters applied
        
        Args:
            workorders: List of work order dictionaries
            filters: Optional filter parameters (status, priority, date_range)
            format: Export format (XLSX or CSV)
            
        Returns:
            (file_bytes, content_type)
        """
        # Apply filters if provided (filtering is typically done at DB level,
        # but this allows post-processing if needed)
        filtered_workorders = workorders
        
        if filters:
            status = filters.get('status')
            priority = filters.get('priority')
            
            if status:
                filtered_workorders = [
                    wo for wo in filtered_workorders 
                    if wo.get('status') == status
                ]
            
            if priority:
                filtered_workorders = [
                    wo for wo in filtered_workorders 
                    if wo.get('priority') == priority
                ]
        
        filename = f"workorders_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        return self.export(filtered_workorders, filename, format)
    
    def export_single(
        self,
        workorder: Dict[str, Any],
        format: ExportFormat = ExportFormat.XLSX
    ) -> Tuple[bytes, str]:
        """
        Export a single work order
        
        Args:
            workorder: Single work order dictionary
            format: Export format (XLSX or CSV)
            
        Returns:
            (file_bytes, content_type)
        """
        filename = f"workorder_{workorder.get('order_no', 'single')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        return self.export([workorder], filename, format)
    
    def _export_xlsx(
        self,
        workorders: List[Dict[str, Any]],
        filename: str,
        include_columns: List[str] = None,
        title: str = '工单导出'
    ) -> Tuple[bytes, str]:
        """Export to XLSX format"""
        if EXCEL_ENGINE == 'openpyxl':
            return self._export_openpyxl(workorders, filename, include_columns, title)
        elif EXCEL_ENGINE == 'xlsxwriter':
            return self._export_xlsxwriter(workorders, filename, include_columns, title)
        else:
            raise ImportError("Neither openpyxl nor xlsxwriter is available")
    
    def _export_openpyxl(
        self,
        workorders: List[Dict[str, Any]],
        filename: str,
        include_columns: List[str] = None,
        title: str = '工单导出'
    ) -> Tuple[bytes, str]:
        """Export using openpyxl"""
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Sheet name max 31 chars
        
        # Determine which columns to include
        columns = self.COLUMNS
        if include_columns:
            columns = [c for c in self.COLUMNS if c['key'] in include_columns]
        
        # Write header row
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col['header'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data rows
        for row_idx, wo in enumerate(workorders, 2):
            for col_idx, col in enumerate(columns, 1):
                key = col['key']
                value = wo.get(key)
                
                # Format value
                value = self._format_value(key, value)
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = thin_border
                
                # Apply conditional formatting for priority and status
                if key == 'priority' and value:
                    priority = value.split('-')[0] if '-' in value else value
                    if priority in self.PRIORITY_COLORS:
                        cell.fill = PatternFill(
                            start_color=self.PRIORITY_COLORS[priority],
                            end_color=self.PRIORITY_COLORS[priority],
                            fill_type='solid'
                        )
                
                elif key == 'status' and value:
                    status_key = self._get_status_key(value)
                    if status_key in self.STATUS_COLORS:
                        cell.fill = PatternFill(
                            start_color=self.STATUS_COLORS[status_key],
                            end_color=self.STATUS_COLORS[status_key],
                            fill_type='solid'
                        )
                
                elif key == 'sla_breached' and value:
                    if value == '是':
                        cell.fill = PatternFill(
                            start_color='FF0000',
                            end_color='FF0000',
                            fill_type='solid'
                        )
                        cell.font = Font(color='FFFFFF', bold=True)
        
        # Set column widths
        for col_idx, col in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col['width']
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(workorders) + 1}"
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    def _export_xlsxwriter(
        self,
        workorders: List[Dict[str, Any]],
        filename: str,
        include_columns: List[str] = None,
        title: str = '工单导出'
    ) -> Tuple[bytes, str]:
        """Export using xlsxwriter"""
        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output)
        ws = wb.add_worksheet(title[:31])
        
        # Add formats
        header_format = wb.add_format({
            'bold': True,
            'font_color': 'white',
            'bg_color': '#4472C4',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        cell_format = wb.add_format({
            'align': 'left',
            'valign': 'top',
            'text_wrap': True,
            'border': 1
        })
        
        # Determine columns
        columns = self.COLUMNS
        if include_columns:
            columns = [c for c in self.COLUMNS if c['key'] in include_columns]
        
        # Write header
        for col_idx, col in enumerate(columns):
            ws.write(0, col_idx, col['header'], header_format)
            ws.set_column(col_idx, col_idx, col['width'])
        
        # Write data
        for row_idx, wo in enumerate(workorders, 1):
            for col_idx, col in enumerate(columns):
                key = col['key']
                value = wo.get(key)
                value = self._format_value(key, value)
                ws.write(row_idx, col_idx, value, cell_format)
        
        wb.close()
        output.seek(0)
        
        return output.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    def _export_csv(
        self,
        workorders: List[Dict[str, Any]],
        filename: str,
        include_columns: List[str] = None
    ) -> Tuple[bytes, str]:
        """Export to CSV format"""
        import csv
        
        output = io.StringIO()
        
        # Determine columns
        columns = self.COLUMNS
        if include_columns:
            columns = [c for c in self.COLUMNS if c['key'] in include_columns]
        
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([col['header'] for col in columns])
        
        # Write data
        for wo in workorders:
            row = []
            for col in columns:
                key = col['key']
                value = wo.get(key)
                value = self._format_value(key, value)
                row.append(value)
            writer.writerow(row)
        
        return output.getvalue().encode('utf-8-sig'), 'text/csv; charset=utf-8-sig'
    
    def _format_value(self, key: str, value: Any) -> Any:
        """Format value based on column type"""
        if value is None:
            return ''
        
        if key in ['created_at', 'updated_at', 'expected_end', 'actual_end', 'closed_at']:
            if isinstance(value, datetime):
                return value.strftime('%Y-%m-%d %H:%M:%S')
            elif isinstance(value, str) and value:
                try:
                    dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
                    return dt.strftime('%Y-%m-%d %H:%M:%S')
                except (ValueError, AttributeError):
                    return value
            return value
        
        elif key == 'status':
            return self.STATUS_MAP.get(value, value)
        
        elif key == 'order_type':
            return self.TYPE_MAP.get(value, value)
        
        elif key == 'priority':
            return self.PRIORITY_MAP.get(value, value)
        
        elif key == 'sla_breached':
            return '是' if value else '否'
        
        elif key == 'tags' and isinstance(value, list):
            return ','.join(value)
        
        elif key == 'attachments' and isinstance(value, list):
            return ','.join([a.get('name', str(a)) if isinstance(a, dict) else str(a) for a in value])
        
        elif key == 'description' and len(str(value)) > 32767:
            return str(value)[:32767]
        
        return value
    
    def _get_status_key(self, value: str) -> str:
        """Get status key from Chinese value"""
        for key, val in self.STATUS_MAP.items():
            if val == value:
                return key
        return value
    
    def export_summary_sheet(
        self,
        workorders: List[Dict[str, Any]],
        title: str = '工单汇总'
    ) -> bytes:
        """
        Export summary statistics sheet
        
        Args:
            workorders: List of work order dictionaries
            title: Sheet title
            
        Returns:
            Excel bytes
        """
        if EXCEL_ENGINE != 'openpyxl':
            # For xlsxwriter, return empty bytes
            return b''
        
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]
        
        # Calculate statistics
        total = len(workorders)
        by_status = {}
        by_priority = {}
        by_type = {}
        sla_breached_count = 0
        
        for wo in workorders:
            status = wo.get('status', 'unknown')
            by_status[status] = by_status.get(status, 0) + 1
            
            priority = wo.get('priority', 'unknown')
            by_priority[priority] = by_priority.get(priority, 0) + 1
            
            order_type = wo.get('order_type', 'unknown')
            by_type[order_type] = by_type.get(order_type, 0) + 1
            
            if wo.get('sla_breached'):
                sla_breached_count += 1
        
        # Write title
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        
        # Write statistics
        row = 3
        ws.cell(row=row, column=1, value='统计项目')
        ws.cell(row=row, column=2, value='数量')
        ws.cell(row=row, column=3, value='占比')
        
        header_font = Font(bold=True)
        for col in range(1, 4):
            ws.cell(row=row, column=col).font = header_font
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color='4472C4',
                end_color='4472C4',
                fill_type='solid'
            )
            ws.cell(row=row, column=col).font = Font(bold=True, color='FFFFFFFF')
        
        row += 1
        ws.cell(row=row, column=1, value='工单总数')
        ws.cell(row=row, column=2, value=total)
        ws.cell(row=row, column=3, value='100%')
        
        row += 2
        ws.cell(row=row, column=1, value='按状态统计')
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        for status, count in sorted(by_status.items()):
            ws.cell(row=row, column=1, value=self.STATUS_MAP.get(status, status))
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f'{count/total*100:.1f}%' if total else '0%')
            row += 1
        
        row += 1
        ws.cell(row=row, column=1, value='按优先级统计')
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        for priority, count in sorted(by_priority.items()):
            ws.cell(row=row, column=1, value=self.PRIORITY_MAP.get(priority, priority))
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f'{count/total*100:.1f}%' if total else '0%')
            row += 1
        
        row += 1
        ws.cell(row=row, column=1, value='按类型统计')
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        for order_type, count in sorted(by_type.items()):
            ws.cell(row=row, column=1, value=self.TYPE_MAP.get(order_type, order_type))
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f'{count/total*100:.1f}%' if total else '0%')
            row += 1
        
        row += 2
        ws.cell(row=row, column=1, value='SLA超时')
        ws.cell(row=row, column=2, value=sla_breached_count)
        ws.cell(row=row, column=3, value=f'{sla_breached_count/total*100:.1f}%' if total else '0%')
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()