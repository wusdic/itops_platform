"""
Device Importer Module
批量导入设备模块 - 解析Excel/CSV，验证数据，批量创建设备

支持字段: name, ip_address, device_type, vendor, model, snmp_community, location, idc
支持部分成功(返回成功/失败列表)
"""

import io
import logging
import re
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime
from enum import Enum

logger = logging.getLogger(__name__)

# Try to import openpyxl for Excel support
try:
    from openpyxl import Workbook, load_workbook
    EXCEL_ENGINE = 'openpyxl'
except ImportError:
    EXCEL_ENGINE = None

# Try to import pandas as fallback
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


class ImportFormat(str, Enum):
    """导入文件格式"""
    XLSX = "xlsx"
    CSV = "csv"
    AUTO = "auto"


class DeviceImportResult:
    """设备导入结果"""
    def __init__(self):
        self.success: List[Dict[str, Any]] = []
        self.failed: List[Dict[str, Any]] = []
        self.total: int = 0
        self.errors: List[str] = []
    
    def add_success(self, row: Dict[str, Any], device_id: int = None):
        """添加成功记录"""
        result = row.copy()
        if device_id:
            result['imported_id'] = device_id
        result['import_status'] = 'success'
        self.success.append(result)
    
    def add_failure(self, row: Dict[str, Any], error: str, row_num: int = None):
        """添加失败记录"""
        result = row.copy()
        result['import_status'] = 'failed'
        result['error'] = error
        if row_num:
            result['row'] = row_num
        self.failed.append(result)
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        return {
            'total': self.total,
            'success_count': len(self.success),
            'failed_count': len(self.failed),
            'success': self.success,
            'failed': self.failed,
            'errors': self.errors
        }


class DeviceImporter:
    """
    设备批量导入器
    
    支持Excel(.xlsx)和CSV格式文件导入
    字段映射:
        - name: 设备名称 (必需)
        - ip_address: IP地址 (必需)
        - device_type: 设备类型 (必需)
        - vendor: 厂商 (可选)
        - model: 型号 (可选)
        - snmp_community: SNMP Community (可选)
        - location: 位置 (可选)
        - idc: 机房 (可选)
    """
    
    # 必填字段
    REQUIRED_FIELDS = ['name', 'ip_address', 'device_type']
    
    # 可选字段
    OPTIONAL_FIELDS = ['vendor', 'model', 'snmp_community', 'location', 'idc']
    
    # 所有支持字段
    ALL_FIELDS = REQUIRED_FIELDS + OPTIONAL_FIELDS
    
    # 设备类型映射 (支持中英文)
    DEVICE_TYPE_MAP = {
        # 英文
        'server': 'server_linux',
        'server_linux': 'server_linux',
        'server_windows': 'server_windows',
        'server_vmware': 'server_vmware',
        'server_hyperv': 'server_hyperv',
        'server_kvm': 'server_kvm',
        'network': 'network_switch',
        'network_switch': 'network_switch',
        'network_router': 'network_router',
        'network_firewall': 'network_firewall',
        'network_waf': 'network_waf',
        'network_loadbalancer': 'network_loadbalancer',
        'network_vpn': 'network_vpn',
        'network_ap': 'network_ap',
        'network_ac': 'network_ac',
        'security': 'security_ids',
        'security_ids': 'security_ids',
        'security_ips': 'security_ips',
        'security_antivirus': 'security_antivirus',
        'storage': 'storage_nas',
        'storage_array': 'storage_array',
        'storage_nas': 'storage_nas',
        'storage_tape': 'storage_tape',
        'other': 'other',
        # 中文
        '服务器': 'server_linux',
        'linux服务器': 'server_linux',
        'windows服务器': 'server_windows',
        '交换机': 'network_switch',
        '路由器': 'network_router',
        '防火墙': 'network_firewall',
        'waf': 'network_waf',
        '负载均衡': 'network_loadbalancer',
        '负载均衡器': 'network_loadbalancer',
        'vpn': 'network_vpn',
        '无线ap': 'network_ap',
        '无线控制器': 'network_ac',
        '入侵检测': 'security_ids',
        '入侵防御': 'security_ips',
        '杀毒软件': 'security_antivirus',
        '存储': 'storage_nas',
        '存储阵列': 'storage_array',
        'nas': 'storage_nas',
        '磁带库': 'storage_tape',
        '其他': 'other',
    }
    
    # IP地址正则
    IP_PATTERN = re.compile(
        r'^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$'
    )
    
    def __init__(self):
        """初始化导入器"""
        pass
    
    def parse_file(self, file_content: bytes, filename: str = None, format: ImportFormat = ImportFormat.AUTO) -> List[Dict[str, Any]]:
        """
        解析导入文件
        
        Args:
            file_content: 文件内容(字节)
            filename: 文件名(用于判断格式)
            format: 文件格式
            
        Returns:
            设备数据列表
        """
        if format == ImportFormat.AUTO:
            if filename:
                if filename.lower().endswith('.csv'):
                    format = ImportFormat.CSV
                elif filename.lower().endswith(('.xlsx', '.xls')):
                    format = ImportFormat.XLSX
                else:
                    # Default to CSV
                    format = ImportFormat.CSV
            else:
                format = ImportFormat.CSV
        
        if format == ImportFormat.CSV:
            return self._parse_csv(file_content)
        elif format == ImportFormat.XLSX:
            return self._parse_xlsx(file_content)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    def _parse_csv(self, file_content: bytes) -> List[Dict[str, Any]]:
        """解析CSV文件"""
        try:
            # Try pandas first
            if PANDAS_AVAILABLE:
                df = pd.read_csv(io.BytesIO(file_content), dtype=str, keep_default_na=False)
                df = df.fillna('')
                return df.to_dict('records')
        except Exception as e:
            logger.warning(f"Pandas CSV parsing failed, falling back to manual parsing: {e}")
        
        # Fallback to manual CSV parsing
        content = file_content.decode('utf-8-sig')
        lines = content.strip().split('\n')
        
        if not lines:
            return []
        
        # Parse header
        header_line = lines[0]
        headers = [h.strip().strip('"') for h in header_line.split(',')]
        
        # Normalize headers
        normalized_headers = [self._normalize_header(h) for h in headers]
        
        # Parse data rows
        records = []
        for i, line in enumerate(lines[1:], start=2):
            if not line.strip():
                continue
            
            values = self._parse_csv_line(line)
            if len(values) != len(headers):
                # Handle uneven columns
                values = values + [''] * (len(headers) - len(values))
            
            record = {}
            for j, h in enumerate(normalized_headers):
                if h and j < len(values):
                    record[h] = values[j].strip().strip('"')
            
            if record:
                record['_row_num'] = i
                records.append(record)
        
        return records
    
    def _parse_csv_line(self, line: str) -> List[str]:
        """解析CSV行，处理引号"""
        values = []
        current = ''
        in_quotes = False
        
        for char in line:
            if char == '"':
                in_quotes = not in_quotes
            elif char == ',' and not in_quotes:
                values.append(current)
                current = ''
            else:
                current += char
        
        values.append(current)
        return values
    
    def _parse_xlsx(self, file_content: bytes) -> List[Dict[str, Any]]:
        """解析Excel文件"""
        if not EXCEL_ENGINE:
            raise ImportError("openpyxl is required for Excel parsing. Please install it with: pip install openpyxl")
        
        wb = load_workbook(io.BytesIO(file_content), data_only=True)
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(self._normalize_header(cell.value))
        
        # Parse data rows
        records = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell for cell in row):
                continue
            
            record = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(headers) and headers[col_idx]:
                    record[headers[col_idx]] = str(value) if value is not None else ''
            
            if record:
                record['_row_num'] = row_idx
                records.append(record)
        
        wb.close()
        return records
    
    def _normalize_header(self, header: str) -> str:
        """标准化列名"""
        if not header:
            return ''
        
        header = str(header).strip().lower()
        
        # Direct mapping
        mapping = {
            '设备名称': 'name',
            'name': 'name',
            'hostname': 'name',
            '设备名': 'name',
            'ip': 'ip_address',
            'ip地址': 'ip_address',
            'ip_address': 'ip_address',
            'ipaddress': 'ip_address',
            '设备类型': 'device_type',
            'type': 'device_type',
            'devicetype': 'device_type',
            'device_type': 'device_type',
            '厂商': 'vendor',
            'vendor': 'vendor',
            'manufacturer': 'vendor',
            '品牌': 'vendor',
            '型号': 'model',
            'model': 'model',
            'snmp': 'snmp_community',
            'snmp_community': 'snmp_community',
            'snmpcommunity': 'snmp_community',
            '位置': 'location',
            'location': 'location',
            '机房': 'idc',
            'idc': 'idc',
        }
        
        return mapping.get(header, header)
    
    def validate_row(self, row: Dict[str, Any]) -> Tuple[bool, str]:
        """
        验证单行数据
        
        Returns:
            (is_valid, error_message)
        """
        # Check required fields
        for field in self.REQUIRED_FIELDS:
            value = row.get(field, '').strip() if row.get(field) else ''
            if not value:
                return False, f"缺少必需字段: {field}"
        
        # Validate name
        name = row.get('name', '').strip()
        if len(name) > 128:
            return False, f"设备名称过长(最大128字符): {name[:20]}..."
        
        # Validate IP address
        ip = row.get('ip_address', '').strip()
        if not self.IP_PATTERN.match(ip):
            return False, f"无效的IP地址: {ip}"
        
        # Validate device_type
        device_type = row.get('device_type', '').strip().lower()
        if device_type not in self.DEVICE_TYPE_MAP:
            return False, f"不支持的设备类型: {device_type}"
        
        return True, ''
    
    def validate_data(self, data: List[Dict[str, Any]]) -> DeviceImportResult:
        """
        验证导入数据(不提交)
        
        Args:
            data: 设备数据列表
            
        Returns:
            DeviceImportResult - 验证结果
        """
        result = DeviceImportResult()
        result.total = len(data)
        
        for i, row in enumerate(data, start=1):
            is_valid, error = self.validate_row(row)
            row_num = row.get('_row_num', i)
            
            if is_valid:
                result.add_success(row)
            else:
                result.add_failure(row, error, row_num)
        
        return result
    
    def import_devices(self, data: List[Dict[str, Any]], username: str = 'system') -> DeviceImportResult:
        """
        批量导入设备
        
        Args:
            data: 设备数据列表
            username: 操作用户
            
        Returns:
            DeviceImportResult - 导入结果
        """
        result = DeviceImportResult()
        result.total = len(data)
        
        from modules.foundation.db_models.device import Device, DeviceType, DeviceStatus
        from modules.foundation.db.client import get_db_session
        
        success_count = 0
        fail_count = 0
        
        for row in data:
            row_num = row.get('_row_num', '?')
            is_valid, error = self.validate_row(row)
            
            if not is_valid:
                result.add_failure(row, error, row_num)
                fail_count += 1
                continue
            
            try:
                # Map device_type
                device_type_str = row.get('device_type', '').strip().lower()
                db_device_type = self.DEVICE_TYPE_MAP.get(device_type_str, 'other')
                
                # Get DeviceType enum
                try:
                    device_type_enum = DeviceType(db_device_type)
                except ValueError:
                    device_type_enum = DeviceType.OTHER
                
                with get_db_session() as session:
                    # Check if device with same IP already exists
                    existing = session.query(Device).filter(
                        Device.ip_address == row.get('ip_address', '').strip()
                    ).first()
                    
                    if existing:
                        result.add_failure(row, f"IP {row.get('ip_address')} 已存在 (ID: {existing.id})", row_num)
                        fail_count += 1
                        continue
                    
                    # Create device
                    device = Device(
                        name=row.get('name', '').strip(),
                        hostname=row.get('name', '').strip(),  # Use name as hostname if not specified
                        ip_address=row.get('ip_address', '').strip(),
                        device_type=device_type_enum,
                        vendor=row.get('vendor', '').strip() or None,
                        model=row.get('model', '').strip() or None,
                        snmp_community=row.get('snmp_community', '').strip() or None,
                        location=row.get('location', '').strip() or None,
                        idc=row.get('idc', '').strip() or None,
                        status=DeviceStatus.UNKNOWN,
                        created_by=username,
                        updated_by=username,
                    )
                    
                    session.add(device)
                    session.commit()
                    session.refresh(device)
                    
                    result.add_success(row, device.id)
                    success_count += 1
                    
            except Exception as e:
                logger.error(f"导入设备失败 (row {row_num}): {e}")
                result.add_failure(row, f"导入失败: {str(e)}", row_num)
                fail_count += 1
        
        return result
    
    def generate_template(self, format: ImportFormat = ImportFormat.XLSX) -> Tuple[bytes, str]:
        """
        生成导入模板
        
        Args:
            format: 模板格式
            
        Returns:
            (file_bytes, content_type)
        """
        headers = ['name', 'ip_address', 'device_type', 'vendor', 'model', 'snmp_community', 'location', 'idc']
        header_names_cn = ['设备名称', 'IP地址', '设备类型', '厂商', '型号', 'SNMP Community', '位置', '机房']
        
        # Sample data
        sample_data = [
            ['Web服务器-01', '192.168.1.101', 'server_linux', 'Dell', 'PowerEdge R740', 'public', '机房A-机柜3', 'IDC-A'],
            ['数据库服务器', '192.168.1.102', 'server_linux', 'HP', 'ProLiant DL380', 'public', '机房A-机柜4', 'IDC-A'],
            ['核心交换机', '192.168.1.1', 'network_switch', 'Cisco', 'Catalyst 9300', 'public', '机房A-机柜1', 'IDC-A'],
            ['防火墙', '192.168.1.254', 'network_firewall', '华为', 'USG6550E', 'public', '机房A-机柜1', 'IDC-A'],
            ['存储阵列', '192.168.1.200', 'storage_array', '华为', 'OceanStor 5300', 'public', '机房B-机柜1', 'IDC-B'],
        ]
        
        if format == ImportFormat.CSV:
            return self._generate_template_csv(headers, header_names_cn, sample_data)
        else:
            return self._generate_template_xlsx(headers, header_names_cn, sample_data)
    
    def _generate_template_xlsx(self, headers: List[str], header_names: List[str], sample_data: List[List]) -> Tuple[bytes, str]:
        """生成Excel模板"""
        if not EXCEL_ENGINE:
            raise ImportError("openpyxl is required for Excel export")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "设备导入模板"
        
        # Write headers
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col_idx, name in enumerate(header_names, 1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write sample data
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Set column widths
        widths = [20, 15, 15, 15, 20, 15, 20, 15]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # Save
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    def _generate_template_csv(self, headers: List[str], header_names: List[str], sample_data: List[List]) -> Tuple[bytes, str]:
        """生成CSV模板"""
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(header_names)
        
        # Write sample data
        for row in sample_data:
            writer.writerow(row)
        
        return output.getvalue().encode('utf-8-sig'), 'text/csv; charset=utf-8-sig'
