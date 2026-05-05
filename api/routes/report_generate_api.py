"""
报告生成和下载API路由
提供报告生成、查询、下载等接口
"""

from typing import Optional, List
from datetime import datetime, timedelta
from enum import Enum
import os

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from api.dependencies import get_db, get_current_user, CurrentUser, PaginationParams
from modules.foundation.db_models.base import Base
from modules.foundation.db_models.report_template import (
    ReportTemplate, Report, ReportSchedule, ReportTemplateType, ReportFormat
)
from modules.business.report_generator.generator import (
    ReportGenerator, BUILTIN_TEMPLATES
)


router = APIRouter(tags=["报告生成管理"])


# ============== 枚举定义 ==============

class ReportStatus(str, Enum):
    """报告状态"""
    PENDING = "pending"
    GENERATING = "generating"
    COMPLETED = "completed"
    FAILED = "failed"


# ============== 请求/响应模型 ==============

class ReportGenerateRequest(BaseModel):
    """生成报告请求"""
    template_id: Optional[int] = Field(None, description="模板ID")
    template_type: str = Field(..., description="报告类型: device_status/alert_summary/performance_trend/custom")
    name: str = Field(..., description="报告名称", min_length=1, max_length=200)
    start_date: datetime = Field(..., description="开始日期")
    end_date: datetime = Field(..., description="结束日期")
    format: str = Field("pdf", description="输出格式: pdf/excel")
    filters: Optional[dict] = Field(None, description="筛选条件")
    params: Optional[dict] = Field(None, description="额外参数")
    custom_template: Optional[str] = Field(None, description="自定义模板内容(Jinja2)")


class ReportResponse(BaseModel):
    """报告响应"""
    id: int
    report_no: str
    name: str
    template_type: str
    format: str
    status: str
    start_date: datetime
    end_date: datetime
    file_path: Optional[str]
    file_name: Optional[str]
    file_size: Optional[int]
    created_by: Optional[str]
    created_at: datetime
    completed_at: Optional[datetime]
    
    class Config:
        from_attributes = True


# ============== 辅助函数 ==============

def get_template_content(db: Session, template_id: Optional[int], 
                          template_type: str, custom_template: Optional[str] = None) -> tuple:
    """获取模板内容"""
    if custom_template:
        return custom_template, template_type
    
    if template_id:
        template = db.query(ReportTemplate).filter(ReportTemplate.id == template_id).first()
        if template:
            return template.content, template.template_type
    
    # 使用内置模板
    if template_type in BUILTIN_TEMPLATES:
        return BUILTIN_TEMPLATES[template_type], template_type
    
    raise HTTPException(status_code=400, detail="Template not found")


# ============== API接口 ==============

@router.post("/generate", summary="生成报告")
async def generate_report(
    request: ReportGenerateRequest,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """
    生成报告
    
    - **template_id**: 模板ID(可选)
    - **template_type**: 报告类型(device_status/alert_summary/performance_trend/custom)
    - **name**: 报告名称
    - **start_date**: 开始日期
    - **end_date**: 结束日期
    - **format**: 输出格式(pdf/excel)
    - **filters**: 筛选条件
    - **params**: 额外参数
    - **custom_template**: 自定义模板内容
    """
    try:
        # 获取模板内容
        template_content, actual_type = get_template_content(
            db, request.template_id, request.template_type, request.custom_template
        )
        
        # 生成报告
        generator = ReportGenerator()
        result = generator.generate(
            template_content=template_content,
            template_type=actual_type,
            start_date=request.start_date,
            end_date=request.end_date,
            format=request.format,
            filters=request.filters,
            params={
                "title": request.name,
                "generated_by": current_user.username,
                **(request.params or {})
            }
        )
        
        # 保存报告记录
        report = Report(
            report_no=result["report_no"],
            name=request.name,
            template_id=request.template_id,
            template_type=actual_type,
            start_date=request.start_date,
            end_date=request.end_date,
            format=request.format,
            filters=request.filters,
            params=request.params,
            status="completed",
            file_path=result["file_path"],
            file_name=result["file_name"],
            file_size=result["file_size"],
            created_by=current_user.username,
            completed_at=datetime.now()
        )
        
        db.add(report)
        db.commit()
        db.refresh(report)
        
        return {
            "id": report.id,
            "report_no": report.report_no,
            "name": report.name,
            "status": report.status,
            "file_name": report.file_name,
            "download_url": f"/api/v1/report/reports/{report.id}/download",
            "generated_at": report.completed_at.isoformat()
        }
        
    except Exception as e:
        # 记录失败
        report = Report(
            report_no=f"RPT-FAILED-{datetime.now().strftime('%Y%m%d%H%M%S')}",
            name=request.name,
            template_type=request.template_type,
            start_date=request.start_date,
            end_date=request.end_date,
            format=request.format,
            status="failed",
            error_message=str(e),
            created_by=current_user.username
        )
        db.add(report)
        db.commit()
        
        raise HTTPException(status_code=500, detail=f"Report generation failed: {str(e)}")


@router.get("/", summary="获取报告列表")
async def list_reports(
    template_type: Optional[str] = Query(None, description="报告类型过滤"),
    status: Optional[str] = Query(None, description="状态过滤"),
    format: Optional[str] = Query(None, description="格式过滤"),
    start_date: Optional[datetime] = Query(None, description="开始日期过滤"),
    end_date: Optional[datetime] = Query(None, description="结束日期过滤"),
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    pagination: PaginationParams = Depends(PaginationParams),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """
    获取报告列表
    
    - **template_type**: 按报告类型过滤
    - **status**: 按状态过滤(pending/generating/completed/failed)
    - **format**: 按格式过滤(pdf/excel)
    - **start_date**: 开始日期过滤
    - **end_date**: 结束日期过滤
    - **keyword**: 关键词搜索(名称/报告编号)
    """
    query = db.query(Report)
    
    if template_type:
        query = query.filter(Report.template_type == template_type)
    
    if status:
        query = query.filter(Report.status == status)
    
    if format:
        query = query.filter(Report.format == format)
    
    if start_date:
        query = query.filter(Report.start_date >= start_date)
    
    if end_date:
        query = query.filter(Report.end_date <= end_date)
    
    if keyword:
        query = query.filter(
            Report.name.contains(keyword) | Report.report_no.contains(keyword)
        )
    
    total = query.count()
    items = query.order_by(Report.created_at.desc())\
                 .offset((pagination.page - 1) * pagination.page_size)\
                 .limit(pagination.page_size)\
                 .all()
    
    return {
        "items": [ReportResponse.model_validate(r) for r in items],
        "total": total,
        "page": pagination.page,
        "page_size": pagination.page_size,
    }


@router.get("/{report_id}", summary="获取报告详情")
async def get_report(
    report_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """获取报告的详细信息"""
    report = db.query(Report).filter(Report.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    
    response = ReportResponse.model_validate(report)
    
    # 获取关联的模板信息
    template = None
    if report.template_id:
        template = db.query(ReportTemplate).filter(ReportTemplate.id == report.template_id).first()
    
    return {
        **response.model_dump(),
        "template": {
            "id": template.id,
            "name": template.name,
            "template_type": template.template_type
        } if template else None,
        "report_data": report.report_data
    }


@router.delete("/{report_id}", summary="删除报告")
async def delete_report(
    report_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """删除报告"""
    report = db.query(Report).filter(Report.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    
    # 删除文件
    if report.file_path and os.path.exists(report.file_path):
        try:
            os.remove(report.file_path)
        except Exception as e:
            print(f"Failed to delete file: {e}")
    
    db.delete(report)
    db.commit()
    
    return {"message": "Report deleted successfully"}


@router.get("/{report_id}/download", summary="下载报告")
async def download_report(
    report_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """
    下载报告文件
    
    返回文件流或预签名URL
    """
    report = db.query(Report).filter(Report.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    
    if report.status != "completed":
        raise HTTPException(status_code=400, detail=f"Report status is {report.status}")
    
    if not report.file_path or not os.path.exists(report.file_path):
        raise HTTPException(status_code=404, detail="Report file not found")
    
    # 根据格式设置媒体类型
    media_types = {
        "pdf": "application/pdf",
        "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }
    media_type = media_types.get(report.format, "application/octet-stream")
    
    return FileResponse(
        path=report.file_path,
        filename=report.file_name,
        media_type=media_type
    )


@router.post("/{report_id}/regenerate", summary="重新生成报告")
async def regenerate_report(
    report_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """使用相同的参数重新生成报告"""
    original = db.query(Report).filter(Report.id == report_id).first()
    if not original:
        raise HTTPException(status_code=404, detail="Report not found")
    
    # 获取模板内容
    template_content, actual_type = get_template_content(
        db, original.template_id, original.template_type
    )
    
    # 生成新报告
    generator = ReportGenerator()
    result = generator.generate(
        template_content=template_content,
        template_type=actual_type,
        start_date=original.start_date,
        end_date=original.end_date,
        format=original.format,
        filters=original.filters,
        params=original.params
    )
    
    # 删除旧文件
    if original.file_path and os.path.exists(original.file_path):
        try:
            os.remove(original.file_path)
        except Exception:
            pass
    
    # 更新记录
    original.report_no = result["report_no"]
    original.file_path = result["file_path"]
    original.file_name = result["file_name"]
    original.file_size = result["file_size"]
    original.status = "completed"
    original.completed_at = datetime.now()
    
    db.commit()
    
    return {
        "id": original.id,
        "report_no": original.report_no,
        "status": original.status,
        "download_url": f"/api/v1/report/reports/{original.id}/download"
    }


@router.get("/stats/overview", summary="获取报告统计")
async def get_report_stats(
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """获取报告统计信息"""
    # 统计总数
    total = db.query(Report).count()
    completed = db.query(Report).filter(Report.status == "completed").count()
    failed = db.query(Report).filter(Report.status == "failed").count()
    
    # 按类型统计
    by_type = {}
    for template_type in ["device_status", "alert_summary", "performance_trend", "custom"]:
        count = db.query(Report).filter(Report.template_type == template_type).count()
        by_type[template_type] = count
    
    # 按格式统计
    by_format = {}
    for fmt in ["pdf", "excel"]:
        count = db.query(Report).filter(Report.format == fmt).count()
        by_format[fmt] = count
    
    # 计算总大小
    total_size = db.query(Report).filter(
        Report.status == "completed",
        Report.file_size.isnot(None)
    ).with_entities(func.sum(Report.file_size)).scalar() or 0
    
    # 最近生成
    recent = db.query(Report).order_by(Report.created_at.desc()).limit(5).all()
    
    return {
        "total": total,
        "completed": completed,
        "failed": failed,
        "by_type": by_type,
        "by_format": by_format,
        "total_size": total_size,
        "recent_reports": [ReportResponse.model_validate(r) for r in recent]
    }


@router.get("/export", summary="导出报告列表")
async def export_reports(
    template_type: Optional[str] = Query(None, description="报告类型过滤"),
    start_date: Optional[datetime] = Query(None, description="开始日期"),
    end_date: Optional[datetime] = Query(None, description="结束日期"),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
):
    """
    导出报告列表为Excel
    
    - **template_type**: 报告类型过滤
    - **start_date**: 开始日期
    - **end_date**: 结束日期
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    query = db.query(Report)
    
    if template_type:
        query = query.filter(Report.template_type == template_type)
    if start_date:
        query = query.filter(Report.start_date >= start_date)
    if end_date:
        query = query.filter(Report.end_date <= end_date)
    
    reports = query.order_by(Report.created_at.desc()).all()
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "Reports"
    
    # 标题样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # 写入标题行
    headers = ["ID", "报告编号", "名称", "类型", "格式", "状态", "创建时间", "完成时间"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # 写入数据
    for row, report in enumerate(reports, 2):
        ws.cell(row=row, column=1, value=report.id)
        ws.cell(row=row, column=2, value=report.report_no)
        ws.cell(row=row, column=3, value=report.name)
        ws.cell(row=row, column=4, value=report.template_type)
        ws.cell(row=row, column=5, value=report.format)
        ws.cell(row=row, column=6, value=report.status)
        ws.cell(row=row, column=7, value=report.created_at.strftime("%Y-%m-%d %H:%M:%S") if report.created_at else "")
        ws.cell(row=row, column=8, value=report.completed_at.strftime("%Y-%m-%d %H:%M:%S") if report.completed_at else "")
    
    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    # 保存
    filename = f"reports_export_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    filepath = f"/tmp/{filename}"
    wb.save(filepath)
    
    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )